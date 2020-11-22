import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from enum import Enum
#requests, bs, openpyxl(excel lib) imported 

def dataExtr():
    createXlsx()
    productNumber=0
    for i in range(1,51):
        url="https://www.n11.com/bilgisayar/dizustu-bilgisayar?pg="+str(i)
        #get request then parse
        try:
            r=requests.get(url)
            print(str(r.status_code)+" - data is being taken")
        except Exception:
            print(str(r.status_code)+" ERROR")
        
        bs=BeautifulSoup(r.content,"lxml")
        products=bs.find_all("li",attrs={"class":"column"})

        for product in products:
            propNameL=[]
            propValL=[]
            productName=product.a.get("title")
            productUrl=product.a.get("href")
            productPrice=product.find("ins").next.strip()

            propNameL.append("Name")
            propValL.append(productName)    #add property name then add property value
            propNameL.append("Price")
            propValL.append(productPrice)

            try:
                preq=requests.get(productUrl)
            except Exception:
                print("not recived the product's properties")
            pbs=BeautifulSoup(preq.content,"lxml")
            
            prdProps =pbs.find_all("div",attrs={"class","feaItem"})
            for prdprop in prdProps:
                name=prdprop.find("span",attrs={"class","label"}).text
                try:
                    val=prdprop.find("span",attrs={"class","data"}).text
                except Exception:
                    val=prdprop.find("a",attrs={"class","data"}).find("span").text

                propNameL.append(name)
                propValL.append(val)
            
            saveToXlsx(propNameL,propValL,productNumber+2)
            productNumber+=1
    
        print("{}. page was completed".format(i))


def createXlsx():
    wb=Workbook()   #book
    ws=wb.active    #workstation
    ws.title="datas" #stationNamed

    rows='BCDEFGHIJKLM'
    ws.column_dimensions['A'].width=50
    for i in range(len(rows)):
        ws.column_dimensions[rows[i]].width=20
    
    ws.cell(1,1,"NAME")
    ws.cell(1,2,"Marka")
    ws.cell(1,3,"İşlemci Modeli")
    ws.cell(1,4,"İşletim Sistemi")
    ws.cell(1,5,"SSD")
    ws.cell(1,6,"Disk Kapasitesi")
    ws.cell(1,7,"Sistem Belleği (Gb)")
    ws.cell(1,8,"Ekran Kartı Belleği")
    ws.cell(1,9,"Ekran Boyutu")
    ws.cell(1,10,"Maksimum Ekran Çözünürlük")
    ws.cell(1,11,"Usb 3,0 Desteği")
    ws.cell(1,12,"TurboBoost")
    ws.cell(1,13,"PRICE")

    wb.save("n11_bot.xlsx")   #save as
    wb.close()

def saveToXlsx(nameList,valList,row):
    wb=load_workbook("n11_bot.xlsx")
    ws=wb.active
    vl="DENEMEEE"
    ws.cell(2,3,vl)
    

    for i in range(len(nameList)):
        for j in range(1,ws.max_column+1):
            if nameList[i].lower() in ws.cell(1,j).value.lower():
                ws.cell(row,j,valList[i])


    wb.save("n11_bot.xlsx")
    wb.close()

dataExtr()
